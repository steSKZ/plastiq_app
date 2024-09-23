import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import datetime

## Basic variables
# Save data to excel file
file_path = 'content/plastiq_input_information.xlsx'
df_header = "Kontaktinformationen"

## Functions
# Function to collect contact data
def collect_contact():
    with st.form(key="contact_form"):

        # Input fields for the contact
        nachname = st.text_input("Nachname", key="input_nachname")
        vorname = st.text_input("Vorname", key="input_vorname")
        unternehmen = st.text_input("Unternehmen", key="input_unternehmen")
        position = st.text_input("Position im Unternehmen", key="input_position")
        mail = st.text_input("E-Mail Adresse", key="input_mail")
        telefonnummer = st.text_input("Telefonnummer (z. B. +49 15X XXXXXX)", key="input_telefonnummer")

        # Additional variables for the DataFrame
        id_contact = 1 #to be specified
        timestamp = datetime.datetime.now().timestamp()
        utc_time = datetime.datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")

        # Form submit button
        submit_button = st.form_submit_button(label="Speichern", on_click=update_keys)

        # Store the data in a DataFrame if the form is submitted
        if submit_button:
            contact_data = {
                "ID_Kontakt": [id_contact],
                "Zeit_UTC": [utc_time],
                "Nachname": [nachname],
                "Vorname": [vorname],
                "Unternehmen": [unternehmen],
                "Position": [position],
                "Mail": [mail],
                "Telefonnummer": [telefonnummer],
            }

            contact_df = pd.DataFrame(contact_data)
            return contact_df
    return None

# Function to update dict with session state keys after every submission of form
def update_keys():
    for k in st.session_state.key_dict_contact:
        st.session_state.key_dict_contact[k] = st.session_state[k]

# Function to append data to an existing Excel file
def append_df_to_excel(file_path, df, sheet_name='contact_data', startrow=None, **to_excel_kwargs):
    try:
        # Try to open an existing workbook
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the last row in the existing Excel sheet if startrow is not defined
            if startrow is None:
                writer.workbook = load_workbook(file_path)
                if sheet_name in writer.workbook.sheetnames:
                    startrow = writer.workbook[sheet_name].max_row
                else:
                    startrow = 0

            # Write out the new data below the existing data
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False, **to_excel_kwargs)

    except FileNotFoundError:
        # If the file does not exist, create it with the DataFrame
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, **to_excel_kwargs)

# Function to show dataframe
def show_dataframe(df_header, contact_df):
    st.write(df_header)
    st.dataframe(contact_df)
    
## Excecution of streamlit page
# Initialize keys for contact input form if not available 
if "key_dict_contact" not in st.session_state:
    st.session_state.key_dict_contact = {"input_nachname":"",
                                         "input_vorname":"",
                                         "input_unternehmen":"",
                                         "input_position":"",
                                         "input_mail":"",
                                         "input_telefonnummer":"",
                                         }
# For loop: Create session state key for every key in key_dict_contact
for k in st.session_state.key_dict_contact:
    st.session_state[k] = st.session_state.key_dict_contact[k]

# Streamlit app
st.title("Kontakt")

# Collect contact using the function
contact_df = collect_contact()

# Display and store the dataframe if not None
if contact_df is not None:
    # Append the DataFrame to the existing Excel file
    append_df_to_excel(file_path, contact_df)
    show_dataframe (df_header, contact_df)

# Display buttons to switch between input pages
button_next = st.button("Weiter")
if button_next:
    st.switch_page("subpages/input/company.py")