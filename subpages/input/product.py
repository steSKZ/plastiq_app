import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import datetime

## Basic variables
# Save data to excel file
file_path_input = "content/plastiq_input_information.xlsx"
# Load background data from excel file
file_path_background = "content/background_data_decision_tree.xlsx"
df_header = "Produktinformationen"
# Create list with all relevant materials
wertstoff_typ_list = pd.read_excel(file_path_background, sheet_name = "list_material")["abbreviation"].tolist()

## Functions
# Function to update dict with session state keys after every submission of form
def update_keys():
    for k in st.session_state.key_dict_product:
        st.session_state.key_dict_product[k] = st.session_state[k]

# Function to collect product data
def collect_product(product_number):
    with st.form(key="waste_fraction_form"):
        values_wertstoff_typ = []
        values_wertstoff_name = []
        values_wertstoff_anteil = []
        
        # Create three columns inside the form with captions
        left_column_wertstoff, middle_column_wertstoff, right_column_wertstoff = st.columns([1,1,.6])
        left_column_wertstoff.subheader("Wertstofftyp")
        middle_column_wertstoff.subheader("Bezeichnung")
        right_column_wertstoff.subheader("Anteil (%)")


        for i in range(waste_fractions_number):
                    
            wertstoff_typ = left_column_wertstoff.selectbox(label=f"Wertstofftyp_{i+1}", options=wertstoff_typ_list, key=f"input_wertstoff_typ_{i}", label_visibility="collapsed")
            wertstoff_name = middle_column_wertstoff.text_input(label=f"Wertstoffname_{i+1}", key=f"input_wertstoff_name_{i}", label_visibility="collapsed")
            wertstoff_anteil = right_column_wertstoff.number_input(label=f"Wertstoffanteil_{i+1}", min_value=0.00, max_value=100.00, key=f"input_wertstoff_anteil_{i}", format="%.2f", label_visibility="collapsed")
            values_wertstoff_typ.append(wertstoff_typ)
            values_wertstoff_name.append(wertstoff_name)
            values_wertstoff_anteil.append(wertstoff_anteil)

        # Additional variables for the DataFrame
        id_product = 1 #to be specified
        timestamp = datetime.datetime.now().timestamp()
        utc_time = datetime.datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")

        # Form submit button
        submit_button_product = st.form_submit_button(label='Speichern', on_click=update_keys)

        # Store the data in a DataFrame if the form is submitted
        if submit_button_product:

            product_fractions = {
                "ID_Wertstoff": [id_product],
                "Zeit_UTC": [utc_time],
                "Wertstofftyp": [values_wertstoff_typ],
                "Wertstoffname": [values_wertstoff_name],
                "Wertstoffanteil": [values_wertstoff_anteil],
            }

            product_df = pd.DataFrame(product_fractions)
            return product_df
    return None

# Function to show dataframe
def show_dataframe(header, df):
    st.write(header)
    st.dataframe(df)

# Function to append data to an existing Excel file
def append_df_to_excel(file_path_input, df, sheet_name='product_fractions', startrow=None, **to_excel_kwargs):
    try:
        # Try to open an existing workbook
        with pd.ExcelWriter(file_path_input, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the last row in the existing Excel sheet if startrow is not defined
            if startrow is None:
                writer.workbook = load_workbook(file_path_input)
                if sheet_name in writer.workbook.sheetnames:
                    startrow = writer.workbook[sheet_name].max_row
                else:
                    startrow = 0

            # Write out the new data below the existing data
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False, **to_excel_kwargs)

    except FileNotFoundError:
        # If the file does not exist, create it with the DataFrame
        with pd.ExcelWriter(file_path_input, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, **to_excel_kwargs)

# For loop: Create session state key for every key in key_dict_product
for k in st.session_state.key_dict_product:
    st.session_state[k] = st.session_state.key_dict_product[k]

# Streamlit app
st.title("Wertstoffdaten")
st.header("Zusammensetzung des Wertstoffs", divider="red", help="Bitte gebe zu Beginn an, aus wie vielen Fraktionen der Abfall besteht. Anschließend kannst du die Fraktionen spezifizieren mit dem Materialtyp, der genauen Bezeichnung und dem Anteil am Abfall.")

waste_fractions_number = st.slider(label="Anzahl der verwertbaren Fraktionen im Abfall:", min_value=1, max_value=4, key="input_waste_fraction_number", help="Bitte wähle die Anzahl der einzelnene Wertstoffe aus, aus denen der gesammelte Abfall besteht.", on_change=update_keys)

# Collect contact using the function
product_df = collect_product(waste_fractions_number)

# Display the dataframe if not None
if product_df is not None:

    # Append the DataFrame to the existing Excel file
    append_df_to_excel(file_path_input, product_df)
    show_dataframe (df_header, product_df)

# Display buttons to switch between input pages
left_column_bottom, right_column_bottom = st.columns([.13,1])
button_back = left_column_bottom.button("Zurück")
if button_back:
    st.switch_page("subpages/input/company.py")
button_next = right_column_bottom.button("Weiter")
if button_next:
    st.switch_page("subpages/input/product_origin.py")