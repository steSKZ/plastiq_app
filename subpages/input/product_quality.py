import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import datetime

## Basic variables
# Save data to excel file
file_path = "content/plastiq_input_information.xlsx"
df_header = "Wertstoffqualität"

wertstoff_colour_list = ["blau", "braun", "bunt", "divers", "gelb", "grau", "grün", "natur", "rot", "schwarz"]
wertstoff_contaminants_list = ["keine", "gering", "mittel", "hoch", "sehr hoch"]
additiv_typ_list = [ # list with all common additives
    "Antibeschlagmittel", "Antioxidant", "Antistatika", "Biozide", "Chemische Treibmittel",
    "Flammschutzmittel", "Gleitmittel", "Nukleierungsmittel", "Slip-Additiv", "Säurefänger", 
    "UV-Stabilisator", "Verarbeitungshilfsmittel"
]
fuellstoff_typ_list = [ # list with all common fillers
    "Aluminium-/Magnesiumhydroxid", "Bariumsulfat", "Calciumcarbonat", 
    "Feldspat", "Glasfasern", "Glaskugeln", "Glimmer", "Holzmehl", 
    "Kaolin", "Naturfasern", "Ruß", "Silica", "Talk", "Wollastonit"
]

## Functions
# Function to update dict with session state keys after every submission of form
def update_keys():
    for k in st.session_state.key_dict_product_quality:
        st.session_state.key_dict_product_quality[k] = st.session_state[k]

# Function to collect product quality data
def collect_product_quality(value_fraction, value_origin):
    with st.form(key="waste_quality_form"):
    
        wertstoff_reach = st.selectbox(label="REACH-Konformität", options=["Ja", "Nein"], key="input_wertstoff_reach")
        wertstoff_farbe = st.selectbox(label="Farbe des Abfalls", options=wertstoff_colour_list, key="input_wertstoff_colour")
        wertstoff_reinheit = st.number_input(label="Reinheit des Abfalls (in %)", min_value=0.0, max_value=100.0, format="%.1f", key="input_wertstoff_purity")
        wertstoff_verschmutzungsgrad = st.selectbox(label="Verschmutzungsgrad", options=wertstoff_contaminants_list, key="input_wertstoff_contaminants_level")
        wertstoff_verschmutzungsart = st.text_input(label="Verschmutzungsart", key="input_wertstoff_contaminants_type")

        values_additiv_typ = []
        values_fuellstoff_typ = []

        if value_origin == "Post-Industrial (PI)":
        
            # Create two columns inside the form with captions
            left_column_quality, right_column_quality = st.columns([1,1])
            left_column_quality.subheader("Additive je Material")
            right_column_quality.subheader("Füllstoffe je Material")

            for i in range(value_fraction):

                additiv_typ = left_column_quality.multiselect(label=f"Additive für Material {i+1}", options=additiv_typ_list, key=f"input_additiv_typ_{i}", label_visibility="visible")
                fuellstoff_typ = right_column_quality.multiselect(label=f"Füllstoffe für Material {i+1}", options=fuellstoff_typ_list, key=f"input_fuellstoff_typ_{i}", label_visibility="visible")
                
                values_additiv_typ.append(additiv_typ)
                #values_additiv_anteil.append(additiv_anteil)
                values_fuellstoff_typ.append(fuellstoff_typ)
                #values_fuellstoff_anteil.append(fuellstoff_anteil)

        # Additional variables for the DataFrame
        id_product = 1 #to be specified
        timestamp = datetime.datetime.now().timestamp()
        utc_time = datetime.datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")

        # Form submit button
        submit_button_product_quality = st.form_submit_button(label='Speichern', on_click=update_keys)

        # Store the data in a DataFrame if the form is submitted
        if submit_button_product_quality:

            product_quality = {
                "ID_Wertstoff": [id_product],
                "Zeit_UTC": [utc_time],
                "REACH-Konformität": [wertstoff_reach],
                "Farbe": [wertstoff_farbe],
                "Reinheit": [wertstoff_reinheit],
                "Verschmutzungsgrad": [wertstoff_verschmutzungsgrad],
                "Verschmutzungsart": [wertstoff_verschmutzungsart],
                "Additive": [values_additiv_typ],
                "Füllstoffe": [values_fuellstoff_typ],
            }

            product_df = pd.DataFrame(product_quality)
            return product_df
    return None

# Function to show dataframe
def show_dataframe(header, df):
    st.write(header)
    st.dataframe(df)

# Function to append data to an existing Excel file
def append_df_to_excel(file_path, df, sheet_name='product_quality', startrow=None, **to_excel_kwargs):
    try:
        # Try to open an existing workbook
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the last row in the existing Excel sheet if startrow is not defined
            if startrow is None:
                writer.workbook = load_workbook(file_path)
                if sheet_name in writer.workbook.sheetnames:
                    startrow = writer.workbook[sheet_name].max_row
                else:
                    startrow = 0

            # Write out the new data below the existing data
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False, **to_excel_kwargs)

    except FileNotFoundError:
        # If the file does not exist, create it with the DataFrame
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, **to_excel_kwargs)

# Collect contact using the function
value_fraction = st.session_state.key_dict_product["input_waste_fraction_number"]
value_origin = st.session_state.key_dict_product_origin["input_wertstoff_origin"]

# For loop: Create session state key for every key in key_dict_product
for i in st.session_state.key_dict_product_quality:
    st.session_state[i] = st.session_state.key_dict_product_quality[i]

# Streamlit app
st.title("Wertstoffdaten")
st.header("Qualität des Wertstoffs", divider="red", help="Bitte gebe die Informationen zur Qualität des Wertstoffs an.")

# Function to collect product quality data
product_df = collect_product_quality(value_fraction, value_origin)

# Display the dataframe if not None
if product_df is not None:

    # Append the DataFrame to the existing Excel file
    append_df_to_excel(file_path, product_df)
    show_dataframe (df_header, product_df)

# Display buttons to switch between input pages
left_column_bottom, right_column_bottom = st.columns([.13,1])
button_back = left_column_bottom.button("Zurück")
if button_back:
    st.switch_page("subpages/input/product_origin.py")
button_next = right_column_bottom.button("Weiter")
if button_next:
    st.switch_page("subpages/input/product_quality_additive.py")