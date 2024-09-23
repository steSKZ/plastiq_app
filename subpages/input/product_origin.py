import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import datetime

## Basic variables
# Save data to excel file
file_path = "content/plastiq_input_information.xlsx"
df_header = "Herkunftsinformationen"
wertstoff_origin_list = [ # list with the possible origins of the waste
    "Post-Industrial (PI)", "Post-Consumer (PC) – getrennte Sammlung", "Post-Consumer (PC) – gemischte Sammlung"
]
wertstoff_collection_list = [ # list with the possible ways of collection
    "regional", "lokal als Bringsystem", "haushaltsnah als Holsystem", "haushaltsgebunden als Holsystem"
]
wertstoff_collection_help = "Angabe für die Kategorie PC - getrennte Sammlung, nach welcher Sammlungsart der Wertstoff zusammengetragen wurde \n- regional (Letztbesitzer bringt Altprodukt direkt zum Erstbehandler wie z. B. bei Altfahrzeugen üblich) \n- lokal als Bringsystem (Letztbesitzer bringt z. B. Altbatterien zum Einzelhandel), \n- haushaltsnah als Holsystem (Altpapier, Glas, Textilien etc. in Containern) \n- haushaltsgebunden als Holsystem (Restabfalltonne, gelber Sack, Biotonne, blaue Tonne für Altpapier)"

## Functions
# Function to update dict with session state keys after every submission of form
def update_keys():
    for k in st.session_state.key_dict_product_origin:
        st.session_state.key_dict_product_origin[k] = st.session_state[k]

# Function to collect product origin data
def collect_product_origin(value_origin):
    with st.form(key="waste_origin_form"):
    
        wertstoff_verwendung = st.text_input(label="Ursprüngliche Verwendung", key="input_wertstoff_use")

        if value_origin == "Post-Consumer (PC) – getrennte Sammlung":
            wertstoff_sammlung = st.selectbox(label="Sammlungsart", options=wertstoff_collection_list, key="input_wertstoff_collection", help=wertstoff_collection_help)
            wertstoff_abfallcode = st.text_input(label="Abfallcode", max_chars=9, placeholder="XX XX XX*", key="input_wertstoff_code")
        elif value_origin == "Post-Consumer (PC) – gemischte Sammlung":
            wertstoff_sammlung = None
            wertstoff_abfallcode = st.text_input(label="Abfallcode", max_chars=9, placeholder="XX XX XX*", key="input_wertstoff_code")
        else:
            wertstoff_sammlung = None
            wertstoff_abfallcode = None

        # Additional variables for the DataFrame
        id_product = 1 #to be specified
        timestamp = datetime.datetime.now().timestamp()
        utc_time = datetime.datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")

        # Form submit button
        submit_button_product_origin = st.form_submit_button(label='Speichern', on_click=update_keys)

        # Store the data in a DataFrame if the form is submitted
        if submit_button_product_origin:

            product_origin = {
                "ID_Wertstoff": [id_product],
                "Zeit_UTC": [utc_time],
                "Herkunftskategorie": [value_origin],
                "Ursprüngliche Verwendung": [wertstoff_verwendung],
                "Sammlungsart": [wertstoff_sammlung],
                "Abfallcode": [wertstoff_abfallcode],
            }

            product_df = pd.DataFrame(product_origin)
            return product_df
    return None

# Function to show dataframe
def show_dataframe(header, df):
    st.write(header)
    st.dataframe(df)

# Function to append data to an existing Excel file
def append_df_to_excel(file_path, df, sheet_name='product_origin', startrow=None, **to_excel_kwargs):
    try:
        # Try to open an existing workbook
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the last row in the existing Excel sheet if startrow is not defined
            if startrow is None:
                writer.workbook = load_workbook(file_path)
                if sheet_name in writer.workbook.sheetnames:
                    startrow = writer.workbook[sheet_name].max_row
                else:
                    startrow = 0

            # Write out the new data below the existing data
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False, **to_excel_kwargs)

    except FileNotFoundError:
        # If the file does not exist, create it with the DataFrame
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, **to_excel_kwargs)

    
# For loop: Create session state key for every key in key_dict_product
for k in st.session_state.key_dict_product_origin:
    st.session_state[k] = st.session_state.key_dict_product_origin[k]

# Streamlit app
st.title("Wertstoffdaten")
st.header("Herkunft des Wertstoffs", divider="red", help="Bitte gebe die Informationen zur Herkunft, der Sammlungsart und zur ursprünglichen Verwendung des Wertstoffs an.")

wertstoff_herkunft = st.selectbox(label="Herkunftskategorie", options=wertstoff_origin_list, key="input_wertstoff_origin", on_change=update_keys)

# Collect contact using the function
product_df = collect_product_origin(wertstoff_herkunft)

# Display the dataframe if not None
if product_df is not None:

    # Append the DataFrame to the existing Excel file
    append_df_to_excel(file_path, product_df)
    show_dataframe (df_header, product_df)

# Display buttons to switch between input pages
left_column_bottom, right_column_bottom = st.columns([.13,1])
button_back = left_column_bottom.button("Zurück")
if button_back:
    st.switch_page("subpages/input/product.py")
button_next = right_column_bottom.button("Weiter")
if button_next:
    st.switch_page("subpages/input/product_quality.py")