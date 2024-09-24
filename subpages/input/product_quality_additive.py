import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import datetime

## Basic variables
# Save data to excel file
file_path = "content/plastiq_input_information.xlsx"
df_header = "Zusätze"

## Functions
# Function to update dict with session state keys after every submission of form
def update_keys():
    for k in st.session_state.key_dict_additive_quality:
        st.session_state.key_dict_additive_quality[k] = st.session_state[k]

# Function to collect product quality data
def collect_additive_quality(list_wertstoff_name, list_additiv_typ, list_fuellstoff_typ):
    with st.form(key="waste_additive_form"):
    
        list_additiv_anteil = []
        list_fuellstoff_anteil = []
        list_additiv_anteil_kurz = []
        list_fuellstoff_anteil_kurz = []

        for k in range(len(list_wertstoff_name)):
            
            st.subheader(f"Angaben für Wertstoff {list_wertstoff_name[k]} ")
            left_column_additiv, right_column_additiv = st.columns([1,.5])
            left_column_fuellstoff, right_column_fuellstoff = st.columns([1,.5])


            if list_additiv_typ[k]:
                list_additiv_anteil_kurz = []
                left_column_additiv.write("Additive des Wertstoffs")
                right_column_additiv.write("Anteil in %")

                for l in range(len(list_additiv_typ[k])):
                    left_column_additiv.write(list_additiv_typ[k][l])
                    additiv_anteil = right_column_additiv.number_input(label=f"Additiv_{l+1}_Wertstoff_{k+1}", min_value=0.0, max_value=100.0, format="%.2f", key=f"input_anteil_additiv_{l}_werkstoff_{k}", label_visibility="collapsed")

                    list_additiv_anteil_kurz.append(additiv_anteil)


            if list_fuellstoff_typ[k]:
                list_fuellstoff_anteil_kurz = []
                left_column_fuellstoff.write("Füllstoffe des Wertstoffs")
                right_column_fuellstoff.write("Anteil in %")

                for l in range(len(list_fuellstoff_typ[k])):
                    left_column_fuellstoff.write(list_fuellstoff_typ[k][l])
                    fuellstoff_anteil = right_column_fuellstoff.number_input(label=f"Füllstoff_{l+1}_Wertstoff_{k+1}", min_value=0.0, max_value=100.0, format="%.2f", key=f"input_anteil_fuellstoff_{l}_werkstoff_{k}", label_visibility="collapsed")

                    list_fuellstoff_anteil_kurz.append(fuellstoff_anteil)

            list_additiv_anteil.append(list_additiv_anteil_kurz)
            list_fuellstoff_anteil.append(list_fuellstoff_anteil_kurz)

        # Additional variables for the DataFrame
        id_product = 1 #to be specified
        timestamp = datetime.datetime.now().timestamp()
        utc_time = datetime.datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")

        # Form submit button
        submit_button_additive_quality = st.form_submit_button(label='Speichern', on_click=update_keys)

        # Store the data in a DataFrame if the form is submitted
        if submit_button_additive_quality:

            additive_quality = {
                "ID_Wertstoff": [id_product],
                "Zeit_UTC": [utc_time],
                "Additive": [list_additiv_typ],
                "Additivanteil": [list_additiv_anteil],
                "Füllstoffe": [list_fuellstoff_typ],
                "Füllstoffanteil": [list_fuellstoff_anteil]
            }

            product_df = pd.DataFrame(additive_quality)
            return product_df
    return None

# Function to show dataframe
def show_dataframe(header, df):
    st.write(header)
    st.dataframe(df)

# Function to append data to an existing Excel file
def append_df_to_excel(file_path, df, sheet_name='additive_quality', startrow=None, **to_excel_kwargs):
    try:
        # Try to open an existing workbook
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the last row in the existing Excel sheet if startrow is not defined
            if startrow is None:
                writer.workbook = load_workbook(file_path)
                if sheet_name in writer.workbook.sheetnames:
                    startrow = writer.workbook[sheet_name].max_row
                else:
                    startrow = 0

            # Write out the new data below the existing data
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False, **to_excel_kwargs)

    except FileNotFoundError:
        # If the file does not exist, create it with the DataFrame
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, **to_excel_kwargs)

# Collect product quality data using the function
value_fraction = st.session_state.key_dict_product["input_waste_fraction_number"]
list_wertstoff_name = []
list_additiv_typ = []
list_fuellstoff_typ = []

# Collect all additive and filler information in separate lists
for k in range(value_fraction):
    list_additiv_typ.append(st.session_state.key_dict_product_quality[f"input_additiv_typ_{k}"])
    list_fuellstoff_typ.append(st.session_state.key_dict_product_quality[f"input_fuellstoff_typ_{k}"])
    list_wertstoff_name.append(st.session_state.key_dict_product[f"input_wertstoff_name_{k}"])

# Initialize keys for product input form if not available 
if "key_dict_additive_quality" not in st.session_state:

    st.session_state.key_dict_additive_quality = {}

    for k in range(value_fraction):

        for l in range(len(list_additiv_typ[k])):
            st.session_state.key_dict_additive_quality[f"input_anteil_additiv_{l}_werkstoff_{k}"] = 0.00

        for l in range(len(list_fuellstoff_typ[k])):
            st.session_state.key_dict_additive_quality[f"input_anteil_fuellstoff_{l}_werkstoff_{k}"] = 0.00
    
# For loop: Create session state key for every key in key_dict_product
for k in st.session_state.key_dict_additive_quality:
    st.session_state[k] = st.session_state.key_dict_additive_quality[k]

# Streamlit app
st.title("Wertstoffdaten")
st.header("Additive und Füllstoffe", divider="red", help="Bitte gebe die Informationen zum Anteil der Additive und Füllstoffe in den Materialien an.")

product_df = collect_additive_quality(list_wertstoff_name, list_additiv_typ, list_fuellstoff_typ)

# Display the dataframe if not None
if product_df is not None:

    # Append the DataFrame to the existing Excel file
    append_df_to_excel(file_path, product_df)
    show_dataframe (df_header, product_df)

# Display buttons to switch between input pages
left_column_bottom, right_column_bottom = st.columns([.13,1])
button_back = left_column_bottom.button("Zurück")
if button_back:
    st.switch_page("subpages/input/product_quality.py")
button_next = right_column_bottom.button("Weiter")
if button_next:
    st.switch_page("subpages/input/product_further.py")