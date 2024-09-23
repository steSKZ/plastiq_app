import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import datetime

## Basic variables
# Save data to excel file
file_path = "content/plastiq_input_information.xlsx"
df_header = "Weitere Informationen"
haeufigkeit_turnus_list = ["Tag", "Woche", "Monat", "Quartal", "Jahr"]

## Functions
# Function to update dict with session state keys after every submission of form
def update_keys():
    for k in st.session_state.key_dict_product_amount:
        st.session_state.key_dict_product_amount[k] = st.session_state[k]

# Function to collect product origin data
def collect_product_amount():
    with st.form(key="waste_amount_form"):
        

        wertstoff_menge = st.number_input(label="Verfügbare Menge aktuell (in t)", min_value=0.00, format="%.2f", key="input_wertstoff_menge")
        
        st.subheader("Häufigkeit des anfallenden Wertstoffs")
        left_column_amount, middle_column_amount, right_column_amount= st.columns([.5,.5,1])
        wertstoff_haeufigkeit_menge = left_column_amount.number_input(label="Menge (in t)", min_value=0.00, format="%.2f", key="input_haeufigkeit_menge")
        middle_column_amount.write("pro")
        wertstoff_haeufigkeit_turnus = right_column_amount.selectbox(label="Turnus", options=haeufigkeit_turnus_list, key="input_haeufigkeit_turnus")
        wertstoff_bilder = st.file_uploader(label="Bilder des Wertstoffs hochladen.")
        wertstoff_beschreibung = st.text_area(label="Weitere Beschreibung (freies Feld)", max_chars=400, key="input_wertstoff_beschreibung")

        # Additional variables for the DataFrame
        id_product = 1 #to be specified
        timestamp = datetime.datetime.now().timestamp()
        utc_time = datetime.datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")

        # Form submit button
        submit_button_product_amount = st.form_submit_button(label='Speichern', on_click=update_keys)

        # Store the data in a DataFrame if the form is submitted
        if submit_button_product_amount:

            product_amount = {
                "ID_Wertstoff": [id_product],
                "Zeit_UTC": [utc_time],
                "Menge": [wertstoff_menge],
                "Menge pro Lieferung": [wertstoff_haeufigkeit_menge],
                "Turnus": [wertstoff_haeufigkeit_turnus],
                "Weitere Beschreibung": [wertstoff_beschreibung]
            }

            product_df = pd.DataFrame(product_amount)
            return product_df
    return None

# Function to show dataframe
def show_dataframe(header, df):
    st.write(header)
    st.dataframe(df)

# Function to append data to an existing Excel file
def append_df_to_excel(file_path, df, sheet_name='product_amount', startrow=None, **to_excel_kwargs):
    try:
        # Try to open an existing workbook
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the last row in the existing Excel sheet if startrow is not defined
            if startrow is None:
                writer.workbook = load_workbook(file_path)
                if sheet_name in writer.workbook.sheetnames:
                    startrow = writer.workbook[sheet_name].max_row
                else:
                    startrow = 0

            # Write out the new data below the existing data
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False, **to_excel_kwargs)

    except FileNotFoundError:
        # If the file does not exist, create it with the DataFrame
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, **to_excel_kwargs)

# Initialize keys for product input form if not available 
if "key_dict_product_amount" not in st.session_state:
    st.session_state.key_dict_product_amount = {"input_wertstoff_menge":0,
                                                "input_haeufigkeit_menge":0,
                                                "input_haeufigkeit_turnus":None,
                                                "input_wertstoff_beschreibung":"",
                                                }
    
# For loop: Create session state key for every key in key_dict_product
for k in st.session_state.key_dict_product_amount:
    st.session_state[k] = st.session_state.key_dict_product_amount[k]

# Streamlit app
st.title("Wertstoffdaten")
st.header("Menge und weitere Angaben", divider="red")

# Collect contact using the function
product_df = collect_product_amount()

# Display the dataframe if not None
if product_df is not None:

    # Append the DataFrame to the existing Excel file
    append_df_to_excel(file_path, product_df)
    show_dataframe (df_header, product_df)

# Display buttons to switch between input pages
left_column_bottom, right_column_bottom = st.columns([.13,1])
button_back = left_column_bottom.button("Zurück")
if button_back:
    st.switch_page("subpages/input/product_quality_additive.py")
button_next = right_column_bottom.button("Eingabe abschließen")
if button_next:
    st.switch_page("subpages/output/score.py")