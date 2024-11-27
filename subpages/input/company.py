import streamlit as st
import pandas as pd
from geopy.geocoders import Nominatim
from openpyxl import load_workbook
import datetime

## Basic variables
# Save data to excel file
file_path = "content/plastiq_input_information.xlsx"
df_header = "Unternehmensinformationen"
bundesland_list = [ #list with all federal states of germany
    "Baden-Württemberg",
    "Bayern",
    "Berlin",
    "Brandenburg",
    "Bremen",
    "Hamburg",
    "Hessen",
    "Mecklenburg-Vorpommern",
    "Niedersachsen",
    "Nordrhein-Westfalen",
    "Rheinland-Pfalz",
    "Saarland",
    "Sachsen",
    "Sachsen-Anhalt",
    "Schleswig-Holstein",
    "Thüringen"
]
land_list = [ # list with all countries
    "Afghanistan", "Ägypten", "Albanien", "Algerien", "Andorra", "Angola", "Antigua und Barbuda", 
    "Äquatorialguinea", "Argentinien", "Armenien", "Aserbaidschan", "Äthiopien", "Australien", 
    "Bahamas", "Bahrain", "Bangladesch", "Barbados", "Belarus", "Belgien", "Belize", "Benin", 
    "Bhutan", "Bolivien", "Bosnien und Herzegowina", "Botswana", "Brasilien", "Brunei", 
    "Bulgarien", "Burkina Faso", "Burundi", "Chile", "China", "Costa Rica", "Dänemark", 
    "Demokratische Republik Kongo", "Deutschland", "Dominica", "Dominikanische Republik", 
    "Dschibuti", "Ecuador", "El Salvador", "Elfenbeinküste", "Eritrea", "Estland", "Eswatini", 
    "Fidschi", "Finnland", "Frankreich", "Gabun", "Gambia", "Georgien", "Ghana", "Grenada", 
    "Griechenland", "Guatemala", "Guinea", "Guinea-Bissau", "Guyana", "Haiti", "Honduras", 
    "Indien", "Indonesien", "Irak", "Iran", "Irland", "Island", "Israel", "Italien", "Jamaika", 
    "Japan", "Jemen", "Jordanien", "Kambodscha", "Kamerun", "Kanada", "Kap Verde", "Kasachstan", 
    "Katar", "Kenia", "Kirgisistan", "Kiribati", "Kolumbien", "Komoren", "Kongo", "Kroatien", 
    "Kuba", "Kuwait", "Laos", "Lesotho", "Lettland", "Libanon", "Liberia", "Libyen", 
    "Liechtenstein", "Litauen", "Luxemburg", "Madagaskar", "Malawi", "Malaysia", "Malediven", 
    "Mali", "Malta", "Marokko", "Marshallinseln", "Mauretanien", "Mauritius", "Mexiko", 
    "Mikronesien", "Moldawien", "Monaco", "Mongolei", "Montenegro", "Mosambik", "Myanmar", 
    "Namibia", "Nauru", "Nepal", "Neuseeland", "Nicaragua", "Niederlande", "Niger", "Nigeria", 
    "Nordkorea", "Nordmazedonien", "Norwegen", "Oman", "Österreich", "Osttimor", "Pakistan", 
    "Palau", "Panama", "Papua-Neuguinea", "Paraguay", "Peru", "Philippinen", "Polen", "Portugal", 
    "Ruanda", "Rumänien", "Russland", "Salomonen", "Sambia", "Samoa", "San Marino", 
    "São Tomé und Príncipe", "Saudi-Arabien", "Schweden", "Schweiz", "Senegal", "Serbien", 
    "Seychellen", "Sierra Leone", "Simbabwe", "Singapur", "Slowakei", "Slowenien", "Somalia", 
    "Spanien", "Sri Lanka", "St. Kitts und Nevis", "St. Lucia", "St. Vincent und die Grenadinen", 
    "Südafrika", "Sudan", "Südsudan", "Suriname", "Syrien", "Tadschikistan", "Tansania", "Thailand", 
    "Togo", "Tonga", "Trinidad und Tobago", "Tschad", "Tschechien", "Tunesien", "Türkei", 
    "Turkmenistan", "Tuvalu", "Uganda", "Ukraine", "Ungarn", "Uruguay", "Usbekistan", 
    "Vanuatu", "Vatikanstadt", "Venezuela", "Vereinigte Arabische Emirate", "Vereinigte Staaten", 
    "Vietnam", "Zentralafrikanische Republik", "Zypern"
]
leistungen_list = [ #list with possible services of recycler
    "Andere",
    "Abfallsammlung",
    "Abfallsortierung",
    "Abfallzerkleinerung",
    "Thermische Verwertung",
    "Reinigung",
    "Trocknen",
    "Einmahlen",
    "Granulierung",
    "Extrusion",
    "Spritzguss",
    "Lagerung",
    "Transport",
    "Entstauben",
]

## Functions
# Function to update dict with session state keys after every submission of form
def update_keys():
    for k in st.session_state.key_dict_company:
        st.session_state.key_dict_company[k] = st.session_state[k]

# Function to get latitude and longitude from address
def geocode_address(address):
    geolocator = Nominatim(user_agent="streamlit-app")
    location = geolocator.geocode(address)
    if location:
        return location.latitude, location.longitude
    else:
        return None, None

# Function to collect contact data
def collect_company():
    with st.form(key="company_form"):

        # Input fields for the contact
        unternahmen_name = st.text_input(label="Unternehmen", key="input_unternehmen", disabled=input_unternehmen_disabled)
        strasse = st.text_input(label="Straße und Hausnummer", key="input_strasse")
        plz = st.text_input(label="Postleitzahl", key="input_plz")
        stadt = st.text_input(label="Stadt", key="input_stadt")
        bundesland = st.selectbox(label="Bundesland", options=bundesland_list, key="input_bundesland", placeholder="Bitte wählen Sie eine Option.")
        land = st.selectbox(label="Land", options=land_list, key="input_land", placeholder="Bitte wählen Sie eine Option.")

        # Additional variables for the DataFrame
        id_company = 1 #to be specified
        timestamp = datetime.datetime.now().timestamp()
        utc_time = datetime.datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")

        if verwerter:
            leistungen = st.multiselect(label="Wählen Sie die Leistungen aus, die Sie anbieten.", options=leistungen_list, key="input_leistungen")
        else:
            leistungen = ""

        if bundesland == None or land != "Deutschland":
            bundesland = ""

        # Form submit button
        submit_button_company = st.form_submit_button(label='Speichern', on_click=update_keys)

        # Store the data in a DataFrame if the form is submitted
        if submit_button_company:

            # Combine the address into a single string
            address = f"{strasse}, {plz} {stadt}, {bundesland}, {land}"

            # Get latitude and longitude from geocode function
            latitude, longitude = geocode_address(address)

            # If conversion from address data to coordinates was successful
            if latitude and longitude:
                st.success(f"Koordinaten: {latitude}, {longitude}")
        
                # Create a dataframe with the coordinates for st.map
                st.session_state.coordinates_data = {
                    'latitude': [latitude],
                    'longitude': [longitude]
                }
                
                # Display the map with the coordinates
                #st.map(pd.DataFrame(coordinates_data))
            else:
                st.session_state.coordinates_data = {
                    'latitude': "",
                    'longitude': ""
                }
                st.error("Die Adresse konnte nicht in Koordinaten umgewandelt werden. Bitte überprüfen Sie die Eingaben.")

            company_data = {
                "ID_Unternehmen": [id_company],
                "Zeit_UTC": [utc_time],
                "Unternehmen_Name": [unternahmen_name],
                "Strasse": [strasse],
                "Postleitzahl": [plz],
                "Stadt": [stadt],
                "Bundesland": [bundesland],
                "Land": [land],
                "Erzeuger": [erzeuger],
                "Verwerter": [verwerter],
                "Leistungen": [leistungen]
            }

            company_df = pd.DataFrame(company_data|st.session_state.coordinates_data)
            return company_df
    return None

# Function to show dataframe
def show_dataframe(header, df):
    st.write(header)
    st.dataframe(df)

# Function to append data to an existing Excel file
def append_df_to_excel(file_path, df, sheet_name='company_data', startrow=None, **to_excel_kwargs):
    try:
        # Try to open an existing workbook
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the last row in the existing Excel sheet if startrow is not defined
            if startrow is None:
                writer.workbook = load_workbook(file_path)
                if sheet_name in writer.workbook.sheetnames:
                    startrow = writer.workbook[sheet_name].max_row
                else:
                    startrow = 0

            # Write out the new data below the existing data
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False, **to_excel_kwargs)

    except FileNotFoundError:
        # If the file does not exist, create it with the DataFrame
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, **to_excel_kwargs)

# Initialize keys for company input form if not available 
if "key_dict_company" not in st.session_state:
    st.session_state.key_dict_company = {"input_unternehmen":"",
                                         "input_strasse":"",
                                         "input_plz":"",
                                         "input_stadt":"",
                                         "input_bundesland":None,
                                         "input_land":None,
                                         "input_erzeuger":False,
                                         "input_verwerter":False,
                                         "input_leistungen":[]
                                         }
    
# For loop: Create session state key for every key in key_dict_company
for k in st.session_state.key_dict_company:
    st.session_state[k] = st.session_state.key_dict_company[k]

# Take company name from input of contact page
if "key_dict_contact" in st.session_state and st.session_state.key_dict_contact["input_unternehmen"] != "":
    st.session_state["input_unternehmen"] = st.session_state.key_dict_contact["input_unternehmen"]
    input_unternehmen_disabled = True
else:
    input_unternehmen_disabled = False

# Streamlit app
st.title("Unternehmen")

left_column_top, right_column_top = st.columns([1,1])
erzeuger = left_column_top.checkbox(label="Kunststofferzeuger/-verarbeiter?", key="input_erzeuger", help="Bitte wählen Sie aus, ob Sie Kunststoffe erzeugen oder verarbeiten. Eine Mehrfachnennung ist möglich.", on_change=update_keys)
verwerter = right_column_top.checkbox(label="Recycler/Verwerter?", key="input_verwerter", help="Bitte wählen Sie aus, ob Sie Kunststoffe recyceln oder anders verwerten. Eine Mehrfachnennung ist möglich.", on_change=update_keys)


# Collect contact using the function
company_df = collect_company()

# Display the dataframe if not None
if company_df is not None:

    # Append the DataFrame to the existing Excel file
    append_df_to_excel(file_path, company_df)
    show_dataframe (df_header, company_df)

# Display buttons to switch between input pages
left_column_bottom, right_column_bottom = st.columns([.13,1])
button_back = left_column_bottom.button("Zurück")
if button_back:
    st.switch_page("subpages/input/contact.py")
button_next = right_column_bottom.button("Weiter")
if button_next:
    st.switch_page("subpages/input/product.py")