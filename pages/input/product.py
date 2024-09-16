import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import datetime

## Basic variables
# Save data to excel file
file_path = "content/plastiq_input_information.xlsx"
df_header = "Produktinformationen"
wertstoff_typ_list = [ # list with all common plastic and metal sorts
    "Kunststoff – ABS", "Kunststoff – ASA", "Kunststoff – Duroplaste", "Kunststoff – Elastomere", 
    "Kunststoff – PA", "Kunststoff – PBT", "Kunststoff – PC", "Kunststoff – PE-HD", 
    "Kunststoff – PE-LD", "Kunststoff – PE-LLD", "Kunststoff – PE-MD", "Kunststoff – PEEK", 
    "Kunststoff – PET", "Kunststoff – PMMA", "Kunststoff – POM", "Kunststoff – PP", 
    "Kunststoff – PS", "Kunststoff – PS-E/XPS", "Kunststoff – PUR", "Kunststoff – PVC", 
    "Kunststoff – SAN", "Kunststoff – Sonstige Thermoplaste", "Metall – Aluminium", 
    "Metall – Blei", "Metall – Chrom", "Metall – Eisen", "Metall – Kupfer", 
    "Metall – Magnesium", "Metall – Nickel", "Metall – Stahl", "Metall – Titan", 
    "Metall – Zink"
]
additiv_typ_list = [ # list with all common additives
    "Antioxidant", "UV-Stabilisator", "Säurefänger", "Gleitmittel", "Verarbeitungshilfesmittel", 
    "Antistatika", "Antibeschlagmittel", "Slip-Additiv", "Nukleierungsmittel", 
    "Chemische Treibmittel", "Flammschutzmittel", "Biozide"
]




## Functions
# Function to update dict with session state keys after every submission of form
def update_keys():
    for k in st.session_state.key_dict_product:
        st.session_state.key_dict_product[k] = st.session_state[k]

# Function to collect product data
def collect_product(product_number):
    with st.form(key="waste_fraction_form"):
        values_wertstoff_typ = []
        values_wertstoff_anteil = []
        values_additiv_typ = []
        values_additiv_anteil = []

        
        for i in range(waste_fractions_number):
            left_column_wertstoff, right_column_wertstoff = st.columns([1,.3])
            
            wertstoff_typ = left_column_wertstoff.selectbox(label=f"Wertstofftyp_{i+1}", options=wertstoff_typ_list, key=f"input_wertstoff_typ_{i}", label_visibility="collapsed")
            wertstoff_anteil = right_column_wertstoff.number_input(label=f"Wertstoffanteil_{i+1}", min_value=0.00, max_value=100.00, key=f"input_wertstoff_anteil_{i}", format="%.2f", label_visibility="collapsed")
            additiv_typ = left_column_wertstoff.selectbox(label=f"Additive_{i+1}", options=additiv_typ_list, key=f"input_additiv_typ_{i}", label_visibility="collapsed")
            additiv_anteil = right_column_wertstoff.number_input(label=f"Additivanteil_{i+1}", min_value=0.00, max_value=100.00, key=f"input_additiv_anteil_{i}", format="%.2f", label_visibility="collapsed")
            values_wertstoff_typ.append(wertstoff_typ)
            values_wertstoff_anteil.append(wertstoff_anteil)
            values_additiv_typ.append(additiv_typ)
            values_additiv_anteil.append(additiv_anteil)

        # Additional variables for the DataFrame
        id_product = 1 #to be specified
        timestamp = datetime.datetime.now().timestamp()
        utc_time = datetime.datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")

        # Form submit button
        submit_button_product = st.form_submit_button(label='Speichern', on_click=update_keys)

        # Store the data in a DataFrame if the form is submitted
        if submit_button_product:

            product_data = {
                "ID_Wertstoff": [id_product],
                "Zeit_UTC": [utc_time],
                "Wertstofftyp": [values_wertstoff_typ],
                "Wertstoffanteil": [values_wertstoff_anteil],
            }

            product_df = pd.DataFrame(product_data)
            return product_df
    return None

# Function to show dataframe
def show_dataframe(header, df):
    st.write(header)
    st.dataframe(df)

# Function to append data to an existing Excel file
def append_df_to_excel(file_path, df, sheet_name='product_data', startrow=None, **to_excel_kwargs):
    try:
        # Try to open an existing workbook
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the last row in the existing Excel sheet if startrow is not defined
            if startrow is None:
                writer.workbook = load_workbook(file_path)
                if sheet_name in writer.workbook.sheetnames:
                    startrow = writer.workbook[sheet_name].max_row
                else:
                    startrow = 0

            # Write out the new data below the existing data
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False, **to_excel_kwargs)

    except FileNotFoundError:
        # If the file does not exist, create it with the DataFrame
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, **to_excel_kwargs)

# Initialize keys for product input form if not available 
if "key_dict_product" not in st.session_state:
    st.session_state.key_dict_product = {"input_waste_fraction_number":1,
                                         "input_wertstoff_typ":None,
                                         }
    
# For loop: Create session state key for every key in key_dict_product
for k in st.session_state.key_dict_product:
    st.session_state[k] = st.session_state.key_dict_product[k]

# Streamlit app
st.title("Abfalldaten")

waste_fractions_number =st.slider(label="Anzahl der verwertbaren Fraktionen im Abfall:", min_value=1, max_value=4, key="input_waste_fraction_number", help="Bitte wähle die Anzahl der einzelnene Wertstoffe aus, aus denen der gesammelte Abfall besteht.", on_change=update_keys)

# Collect contact using the function
product_df = collect_product(waste_fractions_number)

# Display the dataframe if not None
if product_df is not None:

    # Append the DataFrame to the existing Excel file
    append_df_to_excel(file_path, product_df)
    show_dataframe (df_header, product_df)

# Display buttons to switch between input pages
left_column_bottom, right_column_bottom = st.columns([.13,1])
button_back = left_column_bottom.button("Zurück")
if button_back:
    st.switch_page("pages/input/company.py")
button_next = right_column_bottom.button("Auswerten und Ergebnisse ausgeben")
if button_next:
    st.switch_page("pages/output/score.py")